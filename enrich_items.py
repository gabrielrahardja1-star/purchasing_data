"""
enrich_items.py — Enrich item master from monthly material plan Excel files.

IMPORTANT NOTE (future work):
  Item specs (e.g. Φ17.8, 120*12mtr, M16*50) are currently embedded in item names.
  Example: "PC Strand Φ17.8" and "PC Strand Φ15.2" are stored as two separate items.
  This causes item master bloat and makes qty/price comparisons across specs impossible.
  Future improvement: add a `specification` column to the items table and normalise
  item names to be spec-agnostic, with spec stored separately.

Usage:
  python3 enrich_items.py [--dry-run]
"""

import openpyxl
import sqlite3
import os
import re
import sys

DB_PATH = os.path.join(os.path.dirname(__file__), 'db', 'procurement.db')
FOLDERS = [
    '2025年月度材料计划',
]
EXTRA_FILES = [
    '2026年04月生产材料计划 UPDATED VERSION.xlsx',
    '2026年03月面包房食品采购计划·.xlsx',
    '1月.xlsx',
]

DRY_RUN = '--dry-run' in sys.argv

# ── Helpers ────────────────────────────────────────────────────────────────────

def split_cn_en(text):
    """
    Split a mixed Chinese/English name into (name_en, name_cn).
    Pattern: Chinese chars come first, then English. Or vice versa.
    Examples:
      "锚索 PC Strand Φ17.8"  → name_cn="锚索", name_en="PC Strand Φ17.8"
      "Bread Flour"           → name_cn="",     name_en="Bread Flour"
      "PC Strand 锚索"        → name_cn="锚索", name_en="PC Strand"
    """
    text = text.strip()
    # Find all Chinese character spans
    cn_parts = re.findall(r'[\u4e00-\u9fff\u3400-\u4dbf]+', text)
    en_parts = re.sub(r'[\u4e00-\u9fff\u3400-\u4dbf]+', '', text).strip()
    en_parts = re.sub(r'\s+', ' ', en_parts).strip()
    name_cn = ' '.join(cn_parts).strip()
    name_en = en_parts.strip()
    return name_en, name_cn


def guess_category(name_en, name_cn, material_type=''):
    """Rough category mapping from name keywords."""
    combined = (name_en + ' ' + name_cn + ' ' + material_type).lower()
    # Consumables/lubricants first (before food, to avoid "oil" false-match)
    if any(k in combined for k in ['grease', 'lubric', 'hydraulic', 'gear oil', 'engine oil',
                                    'anti-wear', 'total loss', 'coolant', 'solvent', 'paint', 'thinner',
                                    '润滑', '液压油', '齿轮油', '机械油', '黄油', '全损耗', '抗磨',
                                    '防锈', '涂料', '油漆', '稀释']):
        return 'Consumables'
    if any(k in combined for k in ['flour', 'sugar', 'salt', 'rice', 'bread', 'cake',
                                    'spice', 'sauce', 'milk', 'egg', 'meat', 'fish', 'vegetable',
                                    'cooking oil', 'palm oil', 'soy', 'seasoning', 'noodle',
                                    '面粉', '食品', '食材', '面包', '食用', '调料', '食盐', '大米']):
        return 'Food'
    if any(k in combined for k in ['bolt', 'nut', 'screw', 'fastener', 'washer', 'anchor tray',
                                    'rivet', 'clip', 'clamp',
                                    '螺丝', '螺栓', '螺母', '锚具', '托盘', '垫片', '卡扣']):
        return 'Fasteners'
    if any(k in combined for k in ['pipe', 'beam', 'channel', 'plate', 'steel', 'rail', 'mesh',
                                    'strand', 'rebar', 'i-beam', 'h-beam', 'angle iron', 'flat bar',
                                    '钢管', '槽钢', '锚索', '钢板', '工字钢', 'pc strand',
                                    '轨道', '钢丝网', '锚网', '钢筋']):
        return 'Structural'
    if any(k in combined for k in ['cable', 'electric', 'motor', 'battery', 'switch', 'breaker',
                                    'transformer', 'relay', 'fuse', 'inverter', 'sensor', 'lamp',
                                    '电缆', '电线', '电机', '电池', '开关', '变压器', '熔断',
                                    '传感器', '灯', '电气']):
        return 'Electrical'
    if any(k in combined for k in ['pump', 'valve', 'bearing', 'seal', 'belt', 'gear', 'shaft',
                                    'coupling', 'sprocket', 'chain', 'filter', 'cylinder',
                                    '泵', '阀', '轴承', '密封', '皮带', '齿轮', '传动',
                                    '联轴', '链条', '过滤', '油缸']):
        return 'Mechanical Parts'
    if any(k in combined for k in ['helmet', 'glove', 'goggle', 'vest', 'boot', 'mask', 'harness',
                                    'safety', 'ppe', 'respirator', 'earmuff',
                                    '安全', '防护', '手套', '头盔', '口罩', '防尘', '安全帽']):
        return 'Safety'
    if any(k in combined for k in ['cement', 'concrete', 'brick', 'sand', 'timber', 'wood',
                                    'plywood', 'tile', 'insulation', 'sealant',
                                    '水泥', '砖', '木材', '建材', '保温', '密封胶']):
        return 'Building/Civil'
    if any(k in combined for k in ['wrench', 'spanner', 'hammer', 'drill', 'grinder', 'cutter',
                                    'shovel', 'pick', 'chisel', 'saw', 'plier', 'screwdriver',
                                    'tape measure', 'level', 'torch',
                                    '扳手', '锤', '钻', '磨', '锹', '钎', '锯', '钳', '螺丝刀',
                                    '卷尺', '手电']):
        return 'Tools'
    if any(k in combined for k in ['tape', 'rope', 'hose', 'gasket', 'o-ring', 'adhesive',
                                    'cleaning', 'cloth', 'bag', 'container',
                                    '胶布', '绳', '软管', '垫圈', '胶水', '清洁', '布', '袋']):
        return 'Consumables'
    return 'Uncategorized'


def normalize_uom(uom):
    """Normalize Chinese UOM to standard abbreviations."""
    mapping = {
        '吨': 'ton', '根': 'pcs', '个': 'pcs', '片': 'pcs',
        '盘': 'roll', '套': 'set', '桶': 'drum', '包': 'bag',
        '箱': 'box', '卷': 'roll', '块': 'pcs', '米': 'm',
        '公斤': 'kg', '千克': 'kg', '升': 'L', '瓶': 'bottle',
        '袋': 'bag', '张': 'pcs', '条': 'pcs', '只': 'pcs',
        '把': 'pcs', '支': 'pcs', '双': 'pair', '对': 'pair',
        '台': 'unit', '件': 'pcs', '节': 'pcs', '组': 'set',
    }
    return mapping.get(uom.strip(), uom.strip())


# ── Extract items from one worksheet ──────────────────────────────────────────

def extract_from_sheet(ws, fname, shname):
    items = []
    header_row = None
    name_col = spec_col = unit_col = type_col = None

    for ri in range(1, 7):
        row_vals = [str(ws.cell(ri, ci).value or '').strip() for ci in range(1, 20)]
        for ci, v in enumerate(row_vals, 1):
            vl = v.lower()
            if v in ('名称', '物料名称', '品名', '品名\n材料名称') or \
               'ingredient' in vl or ('name' in vl and 'file' not in vl):
                name_col = ci; header_row = ri
            if v in ('规格型号', '规格', '型号') or 'spec' in vl:
                spec_col = ci
            if v in ('单位', 'UOM', 'uom', 'unit', '单位\nUnit'):
                unit_col = ci
            if v in ('物料类型', '类别', 'CATEGORY', 'category', '分类'):
                type_col = ci
        if header_row:
            break

    if not header_row or not name_col:
        return items

    for ri in range(header_row + 1, ws.max_row + 1):
        raw_name = str(ws.cell(ri, name_col).value or '').strip()
        spec     = str(ws.cell(ri, spec_col).value or '').strip() if spec_col else ''
        uom      = str(ws.cell(ri, unit_col).value or '').strip() if unit_col else ''
        mat_type = str(ws.cell(ri, type_col).value or '').strip() if type_col else ''

        if not raw_name or raw_name in ('None', 'nan', '-', '—', '序号', 'NO', 'No'):
            continue
        # Skip rows that look like headers or totals
        if any(raw_name.startswith(k) for k in ('合计', '小计', '总计', '备注', 'Total', 'Note')):
            continue

        # Combine name + spec into full name
        full_name = raw_name
        if spec and spec not in raw_name:
            full_name = f"{raw_name} {spec}"

        name_en, name_cn = split_cn_en(full_name)
        if not name_en and not name_cn:
            continue

        uom = normalize_uom(uom) if uom else ''
        category = guess_category(name_en, name_cn, mat_type)

        items.append({
            'name_en':  name_en,
            'name_cn':  name_cn,
            'uom':      uom,
            'category': category,
            'source':   fname,
        })

    return items


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    conn = sqlite3.connect(DB_PATH)

    # Load existing names for dedup
    existing_en = set()
    existing_cn = set()
    for row in conn.execute('SELECT LOWER(TRIM(name_en)), LOWER(TRIM(name_cn)) FROM items'):
        if row[0]: existing_en.add(row[0])
        if row[1]: existing_cn.add(row[1])

    print(f"Existing items: {conn.execute('SELECT COUNT(*) FROM items').fetchone()[0]}")

    # Collect all Excel files
    all_files = []
    for folder in FOLDERS:
        if os.path.isdir(folder):
            for fname in os.listdir(folder):
                if fname.endswith('.xlsx') and not fname.startswith('~'):
                    all_files.append(os.path.join(folder, fname))
    for f in EXTRA_FILES:
        if os.path.exists(f):
            all_files.append(f)

    # Extract candidates
    seen_names = set()  # dedup within this run
    candidates = []

    for fpath in all_files:
        fname = os.path.basename(fpath)
        try:
            wb = openpyxl.load_workbook(fpath, data_only=True)
            for shname in wb.sheetnames:
                if shname in ('搜索', 'Search', 'Index', 'Sheet3'):
                    continue
                ws = wb[shname]
                for item in extract_from_sheet(ws, fname, shname):
                    name_en_lower = item['name_en'].lower()
                    name_cn_lower = item['name_cn'].lower()
                    # Skip if already in DB
                    if name_en_lower and name_en_lower in existing_en:
                        continue
                    if name_cn_lower and name_cn_lower in existing_cn:
                        continue
                    # Dedup within this run
                    key = name_en_lower or name_cn_lower
                    if key in seen_names:
                        continue
                    seen_names.add(key)
                    candidates.append(item)
        except Exception as e:
            print(f"  Skipped {fname}: {e}")

    print(f"New unique items to add: {len(candidates)}")

    # Show preview by category
    from collections import Counter
    cat_counts = Counter(c['category'] for c in candidates)
    print("\nBy category:")
    for cat, count in sorted(cat_counts.items()):
        print(f"  {cat:20s} {count}")

    if DRY_RUN:
        print("\n-- DRY RUN — no changes written --")
        print("\nSample items:")
        for c in candidates[:20]:
            print(f"  [{c['category']:15s}] {c['name_en'][:45]:45s} | CN: {c['name_cn'][:20]:20s} | {c['uom']}")
        conn.close()
        return

    # Get next item ID
    last_id = conn.execute("SELECT item_id FROM items ORDER BY item_id DESC LIMIT 1").fetchone()
    if last_id:
        next_num = int(last_id[0].replace('ITEM-', '')) + 1
    else:
        next_num = 1

    insert = conn.execute.__self__.prepare if hasattr(conn, 'prepare') else None
    added = 0
    for item in candidates:
        item_id = f"ITEM-{next_num:04d}"
        conn.execute(
            "INSERT INTO items (item_id, name_en, name_cn, category, uom) VALUES (?, ?, ?, ?, ?)",
            (item_id, item['name_en'], item['name_cn'], item['category'], item['uom'])
        )
        next_num += 1
        added += 1

    conn.commit()
    conn.close()
    print(f"\n✓ Added {added} new items to item master.")


if __name__ == '__main__':
    main()
