"""
import_templates.py — Import PR templates from Ci Fanny Database (Rev 1).xlsx into ClickHouse.

Each sheet in the Excel becomes one template. Each row becomes one template line item.
Items are fuzzy-matched to the items table by English name + spec.

Usage:
  python3 import_templates.py --dry-run
  python3 import_templates.py --host 76.13.19.246 --password 'Merge2026!CH'
"""

import sys, os, json, uuid, argparse
from datetime import datetime
from zoneinfo import ZoneInfo

try:
    import requests
except ImportError:
    print("ERROR: pip install requests"); sys.exit(1)
try:
    import openpyxl
except ImportError:
    print("ERROR: pip install openpyxl"); sys.exit(1)

# ── Args ──────────────────────────────────────────────────────────────────────

parser = argparse.ArgumentParser()
parser.add_argument('--dry-run',   action='store_true')
parser.add_argument('--host',      default='localhost')
parser.add_argument('--port',      type=int, default=8123)
parser.add_argument('--db',        default='procurement')
parser.add_argument('--user',      default='procurement_user')
parser.add_argument('--password',  default=os.environ.get('CLICKHOUSE_PASSWORD', 'changeme'))
parser.add_argument('--excel',     default='Final Purchasing Data/Ci Fanny Database (Rev 1).xlsx')
args = parser.parse_args()

DRY_RUN    = args.dry_run
CH_URL     = f"http://{args.host}:{args.port}"
CH_DB      = args.db
CH_AUTH    = (args.user, args.password)
COMPANY_ID = 'PTMMI'
TZ_JKT     = ZoneInfo('Asia/Jakarta')

TEMPLATE_META = {
    'MONTHLY':               {'display_name': 'Monthly Consumables',  'sort_order': 1},
    'REGULARLY':             {'display_name': 'Regular Supplies',      'sort_order': 2},
    'ONCE IN A WHILE':       {'display_name': 'Occasional Items',      'sort_order': 3},
    'LUBE & GREASE':         {'display_name': 'Lubricants & Greases',  'sort_order': 4},
    'TOOLS':                 {'display_name': 'Tools',                  'sort_order': 5},
    'MAINTENANCE SUPPORTING':{'display_name': 'Maintenance Support',   'sort_order': 6},
}

# ── ClickHouse helpers ────────────────────────────────────────────────────────

def ch_insert(table, rows):
    if not rows: return
    ndjson = '\n'.join(json.dumps(r) for r in rows)
    url = f"{CH_URL}/?database={CH_DB}&query=INSERT+INTO+{table}+FORMAT+JSONEachRow"
    resp = requests.post(url, data=ndjson.encode('utf-8'), auth=CH_AUTH,
                         headers={'Content-Type': 'application/x-ndjson'})
    if resp.status_code != 200:
        raise RuntimeError(f"ClickHouse insert into {table} failed: {resp.text[:500]}")

def ch_query(sql):
    resp = requests.post(
        f"{CH_URL}/?database={CH_DB}&default_format=JSONEachRow",
        data=sql.encode('utf-8'), auth=CH_AUTH
    )
    if resp.status_code != 200:
        raise RuntimeError(f"ClickHouse query failed: {resp.text[:500]}")
    lines = [l for l in resp.text.strip().split('\n') if l]
    return [json.loads(l) for l in lines]

def ch_execute(sql):
    resp = requests.post(f"{CH_URL}/?database={CH_DB}", data=sql.encode('utf-8'), auth=CH_AUTH)
    if resp.status_code != 200:
        raise RuntimeError(f"ClickHouse execute failed: {resp.text[:500]}")

def now_ts():
    return datetime.now(TZ_JKT).strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]

def new_uuid():
    return str(uuid.uuid4())

def version():
    return int(datetime.now().timestamp() * 1000)

# ── Text helpers ──────────────────────────────────────────────────────────────

def en_only(s):
    """Extract English part from a bilingual 'CN EN' string."""
    if not s:
        return ''
    s = str(s).strip()
    # Split on first space to separate Chinese chars from English
    # Find first ASCII char that isn't a space
    for i, c in enumerate(s):
        if c.isascii() and c not in (' ', '\t'):
            return s[i:].strip()
    return s.strip()

def cn_only(s):
    """Extract Chinese part from a bilingual 'CN EN' string."""
    if not s:
        return ''
    s = str(s).strip()
    result = []
    for c in s:
        if c.isascii() and c not in (' ', '\t'):
            break
        result.append(c)
    return ''.join(result).strip()

def clean(s):
    if s is None or str(s).strip().lower() == 'none':
        return ''
    return str(s).strip()

# ── Item matching ─────────────────────────────────────────────────────────────

def build_item_index():
    """Fetch all items from ClickHouse, return dict keyed by (name_en_lower, spec_lower)."""
    rows = ch_query("SELECT item_id, name_en, name_cn, spec, uom FROM items FINAL WHERE is_deleted = 0")
    index = {}
    for r in rows:
        key_full = (r['name_en'].strip().lower(), clean(r['spec']).lower())
        key_nospec = (r['name_en'].strip().lower(), '')
        index.setdefault(key_full, r)
        index.setdefault(key_nospec, r)
    return index

def find_item(index, name_en_raw, spec_raw):
    """Try to match a template row to an item_id."""
    name_en = en_only(name_en_raw).strip()
    spec    = clean(spec_raw)

    # Strip Chinese spec prefix (e.g. '安美重载汽车轮毂轴承润滑脂 LG2' → 'LG2')
    spec_en = en_only(spec) if spec else ''

    candidates = [
        (name_en.lower(), spec.lower()),
        (name_en.lower(), spec_en.lower()),
        (name_en.lower(), ''),
    ]
    for key in candidates:
        if key in index:
            return index[key]
    return None

# ── Ensure tables exist ───────────────────────────────────────────────────────

def ensure_tables():
    ch_execute("""
        CREATE TABLE IF NOT EXISTS pr_templates (
            template_id   String,
            company_id    String,
            template_name String DEFAULT '',
            display_name  String DEFAULT '',
            sort_order    UInt8  DEFAULT 0,
            version       UInt64 DEFAULT toUInt64(toUnixTimestamp64Milli(now64(3))),
            is_deleted    UInt8  DEFAULT 0,
            created_at    DateTime64(3, 'Asia/Jakarta') DEFAULT now64(3),
            updated_at    DateTime64(3, 'Asia/Jakarta') DEFAULT now64(3)
        ) ENGINE = ReplacingMergeTree(version)
        ORDER BY (company_id, template_id)
    """)
    ch_execute("""
        CREATE TABLE IF NOT EXISTS pr_template_items (
            template_item_id String,
            company_id       String,
            template_id      String,
            item_id          String DEFAULT '',
            name_en          String DEFAULT '',
            name_cn          String DEFAULT '',
            spec             String DEFAULT '',
            department       String DEFAULT '',
            uom              String DEFAULT '',
            default_qty      Float64 DEFAULT 0,
            sort_order       UInt16  DEFAULT 0,
            version          UInt64  DEFAULT toUInt64(toUnixTimestamp64Milli(now64(3))),
            is_deleted       UInt8   DEFAULT 0,
            created_at       DateTime64(3, 'Asia/Jakarta') DEFAULT now64(3),
            updated_at       DateTime64(3, 'Asia/Jakarta') DEFAULT now64(3)
        ) ENGINE = ReplacingMergeTree(version)
        ORDER BY (company_id, template_item_id)
    """)

# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("PR Template Import: Ci Fanny Database → ClickHouse")
    print("=" * 60)

    if DRY_RUN:
        print("\n[DRY RUN] No data will be written.\n")
    else:
        try:
            resp = requests.get(f"{CH_URL}/ping", auth=CH_AUTH, timeout=5)
            if resp.status_code != 200:
                raise Exception("bad status")
        except Exception:
            print(f"ERROR: Cannot reach ClickHouse at {CH_URL}"); sys.exit(1)
        print(f"Connected to ClickHouse at {CH_URL}\n")
        ensure_tables()

        # Clear existing templates
        ch_execute("TRUNCATE TABLE pr_templates")
        ch_execute("TRUNCATE TABLE pr_template_items")
        print("Cleared existing template data.\n")

    # Load item index for matching
    item_index = {}
    if not DRY_RUN:
        item_index = build_item_index()
        print(f"Loaded {len(item_index)} item index entries.\n")

    # Load Excel
    if not os.path.exists(args.excel):
        print(f"ERROR: Excel file not found: {args.excel}"); sys.exit(1)

    wb = openpyxl.load_workbook(args.excel, read_only=True)
    now = now_ts()
    ver = version()

    total_items = 0
    total_matched = 0

    for sheet_name in wb.sheetnames:
        meta = TEMPLATE_META.get(sheet_name)
        if not meta:
            print(f"  Skipping unknown sheet: {sheet_name}")
            continue

        template_id = new_uuid()
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        template_row = {
            'template_id':   template_id,
            'company_id':    COMPANY_ID,
            'template_name': sheet_name,
            'display_name':  meta['display_name'],
            'sort_order':    meta['sort_order'],
            'version':       ver,
            'is_deleted':    0,
            'created_at':    now,
            'updated_at':    now,
        }

        item_rows = []
        sort_order = 0
        unmatched = []

        for row in rows:
            # Skip blank rows
            if not any(row[:4]):
                continue
            # Handle both 5-col and 6-col sheets
            dept_raw = row[0]
            name_raw = row[1]
            spec_raw = row[2]
            qty_raw  = row[3]
            uom_raw  = row[4] if len(row) > 4 else None

            if not dept_raw or not name_raw:
                continue

            dept_en = en_only(dept_raw)
            name_en = en_only(name_raw)
            name_cn = cn_only(name_raw)
            spec    = clean(spec_raw)
            uom_en  = en_only(uom_raw) if uom_raw else ''
            default_qty = float(qty_raw) if qty_raw and str(qty_raw) not in ('None', '') else 0.0

            # Match to item
            matched = find_item(item_index, name_raw, spec_raw) if item_index else None
            item_id = matched['item_id'] if matched else ''
            if matched:
                total_matched += 1
                # Use DB values if matched
                name_en = matched['name_en'] or name_en
                uom_en  = matched['uom'] or uom_en

            if not item_id:
                unmatched.append(f"{name_en} | {spec}")

            item_rows.append({
                'template_item_id': new_uuid(),
                'company_id':       COMPANY_ID,
                'template_id':      template_id,
                'item_id':          item_id,
                'name_en':          name_en,
                'name_cn':          name_cn,
                'spec':             spec,
                'department':       dept_en,
                'uom':              uom_en,
                'default_qty':      default_qty,
                'sort_order':       sort_order,
                'version':          ver,
                'is_deleted':       0,
                'created_at':       now,
                'updated_at':       now,
            })
            sort_order += 1

        total_items += len(item_rows)
        print(f"  {sheet_name} ({meta['display_name']}): {len(item_rows)} items", end='')
        if item_index:
            matched_count = sum(1 for r in item_rows if r['item_id'])
            print(f"  [{matched_count}/{len(item_rows)} matched to item_id]", end='')
        print()

        if unmatched and not DRY_RUN:
            for u in unmatched:
                print(f"    ⚠ No match: {u}")

        if not DRY_RUN:
            ch_insert('pr_templates', [template_row])
            ch_insert('pr_template_items', item_rows)

    wb.close()

    print(f"\nTotal: {len(wb.sheetnames)} templates, {total_items} items")
    if item_index:
        print(f"Matched to item_id: {total_matched}/{total_items}")

    if DRY_RUN:
        print("\n-- DRY RUN complete — no data written --")
    else:
        print("\n✓ Templates imported successfully.")

if __name__ == '__main__':
    main()
