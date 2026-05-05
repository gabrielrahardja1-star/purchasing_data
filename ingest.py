#!/usr/bin/env python3
"""
ingest.py — Procurement Item Database Ingestion Script

Usage:
    python ingest.py --seed              # seed item_master from the 3 source files
    python ingest.py <file.xlsx>         # ingest a new monthly file
"""

import sys
import csv
import re
import sqlite3
import warnings
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import openpyxl
from rapidfuzz import fuzz, process

warnings.filterwarnings("ignore", category=UserWarning)

# ─── Paths ────────────────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).parent
CSV_PATH = SCRIPT_DIR / "item_master.csv"
DB_PATH = SCRIPT_DIR / "item_master.db"
LOG_PATH = SCRIPT_DIR / "ingest_log.csv"

CSV_COLUMNS = ["item_id", "name_cn", "name_en", "category", "spec", "uom", "department"]
LOG_COLUMNS = ["timestamp", "source_file", "items_parsed", "matched", "new_added", "new_item_ids", "total_after"]

# ─── Category classifier ──────────────────────────────────────────────────────
# Each rule: (category_name, [keywords_to_match_in_combined_name+spec])
# First match wins, so order from most-specific to least-specific.
CATEGORY_RULES = [
    ("Building/Civil", [
        "door", "木门", "门套", "window", "窗户", "tile", "砖", "brick",
        "cement", "水泥", "pipe fitting", "管件", "tank", "水箱",
        "pump house", "magnetic door", "door stopper", "pool tile",
        "vinyl adhesive", "地板", "flooring",
    ]),
    ("Water Treatment", [
        "pac", "氯化铝", "chlorine", "氯", "filter membrane", "滤膜",
        "water pump", "water treatment", "净化", "chemical dosing",
        "coagulant", "flocculant", "activated carbon", "活性炭",
    ]),
    ("Instrumentation", [
        "calibration gas", "标准气", "multimeter", "万用表", "gauge", "仪表",
        "sensor", "传感器", "oscilloscope", "示波器", "monitor", "监测仪",
    ]),
    ("Food", [
        "flour", "面粉", "sugar", "糖", "grain", "谷物", "yeast", "酵母",
        "baking powder", "baking soda", "发酵剂", "乳制品",
        "egg product", "蛋制品", "whole eggs", "fruit", "水果", "nuts", "坚果",
        "spice", "香料", "cinnamon", "espresso", "vanilla", "essence",
        "chocolate", "cocoa", "starch", "淀粉", "bread flour", "cake flour",
        "cereal", "corn flour", "barley flour", "gluten",
        "vinegar", "醋", "sauce", "酱", "cream cheese",
        "milk", "cheese", "unsalted butter", "牛油", "奶油", "食品", "food", "ingredient",
        "大米粉", "玉米粉", "大麦粉", "自发粉", "白砂糖", "葡萄糖", "蔗糖",
        "granulated sugar", "caster sugar", "icing sugar", "brown sugar",
        "all-purpose flour", "whole wheat flour",
        "bread improver", "emulsifier", "whey", "heavy cream", "whipped cream",
        "raisin", "cranberr", "almond", "pistachio", "hazelnut", "walnut",
        "dried fruit", "jam", "jelly", "custard", "gelatin", "agar",
        "baking paper", "parchment", "food coloring", "食用色素",
    ]),
    ("Safety", [
        "safety glasses", "护目镜", "safety helmet", "安全帽", "gloves", "手套",
        "raincoat", "雨衣", "harness", "安全带", "respirator", "mask", "口罩",
        "eyelet", "风筒卯扣", "reflective vest", "safety shoe",
        "welding helmet", "电焊帽", "welding mask", "face shield",
        "hard hat", "first aid", "fire extinguisher", "消防",
    ]),
    ("Structural", [
        "u-channel", "槽钢", "angle steel", "角钢", "h-beam", "工字钢",
        "wire mesh", "锚网", "wiremesh", "seamless pipe", "钢管",
        "flange", "法兰", "pc strand", "锚索", "rail", "轨道",
        "rod bar", "盘圆", "coil bar", "anchor tray", "小托盘",
        "welding rod", "焊条", "steel plate", "钢板",
        "chain-link fencing", "菱形网", "steel belt", "钢带",
        "i-beam", "c-channel", "steel section",
    ]),
    ("Fasteners", [
        "hex bolt", "内六角螺丝", "外六角螺丝", "bolt nut", "螺栓", "螺母",
        "screw extractor", "washer", "垫圈", "anchor bolt", "rivet", "铆钉",
        "wire cable clamp", "绳卡子", "snap ring", "卡簧", "cable clamp",
        "track bolt", "履带螺栓", "sprocket bolt", "驱动齿螺栓",
    ]),
    ("Mechanical Parts", [
        "bearing", "轴承", "rubber gasket", "橡胶垫片", "seal", "密封",
        "check valve", "逆止阀", "ball valve", "球阀", "foot valve", "底阀",
        "gate valve", "闸阀", "tee fitting", "三通", "elbow", "弯头",
        "high pressure hose", "高压软管", "steelwire braided hose",
        "gland packing", "盘根", "graphite gland", "石墨盘根",
        "o-ring", "oring", "nylon rod", "尼龙棒",
        "coupler", "卯扣", "rotor plate", "钢衬板", "sealing plate",
        "filter", "过滤器", "hydraulic", "液压", "strainer", "oil seal", "油封",
        "piston ring", "活塞环", "bushing", "衬套", "liner", "sprocket", "驱动齿",
        "track shoe", "履带板", "track bolt", "muffler", "消声器",
        "pump", "泵", "cylinder", "油缸", "hose", "软管", "fitting", "接头",
        "accumulator", "蓄能器", "relief valve", "安全阀",
        "butterfly valve", "蝶阀", "retaining ring", "u-packing", "wear ring", "lube valve",
        "master pin", "track link", "carrier roller", "track adjuster",
        "gear segment", "water tank", "水箱", "pvc pipe", "pvc管",
        "shoe patch machine", "hand-crank",
    ]),
    ("Electrical", [
        "insulating tape", "绝缘胶布", "cable ties", "扎带",
        "ac contactor", "接触器", "push-button switch", "按钮开关",
        "dual power", "双电源", "battery water", "电瓶水",
        "electrolyte", "电解液", "telephone", "电话机",
        "soldering iron", "电烙铁", "rosin core solder", "焊锡丝",
        "liquid flux", "助焊剂", "spot welder", "点焊机",
        "inverter welding", "逆变焊机", "lithium battery", "锂电池",
        "cable", "电缆", "wire", "电线",
        "lead acid battery", "vrla", "电池", "starter motor", "起动机",
        "alternator", "发电机", "motor", "电机", "generator",
        "contact cleaner", "触点清洁器", "fuse", "保险丝", "circuit breaker", "断路器",
        "led", "light bulb", "灯泡", "single switch", "double switch",
        "heat shrink", "热缩管", "lighting", "照明", "lamp", "灯",
        "德力西", "充电器", "charger",
    ]),
    ("Consumables", [
        "grease", "润滑脂", "lithium grease", "锂基脂", "high temp grease",
        "rust remover", "除锈剂", "paint", "油漆", "spray paint", "自喷漆",
        "transparent tape", "透明胶带", "clear tape", "seal tape", "生料带",
        "welding gas", "oxygen regulator", "氧气表", "acetylene regulator", "乙炔表",
        "gas cutting torch", "割枪", "cutting nozzle", "枪嘴",
        "rag", "碎布", "刷子", "paint brush", "wire brush", "newsprint", "报纸",
        "nylon rope", "尼龙绳", "iron wire", "铁线",
        "mist nozzle", "喷雾头", "high octane fuel", "汽油",
        "slate pen", "石笔", "nylon twine", "风筒线",
        "adhesive", "粘合剂", "sealant", "密封胶", "pvc pipe adhesive", "管胶",
        "silicone", "硅酮", "vinyl adhesive", "地板粘合剂",
        "woven bag", "编织袋", "sand bag", "砂袋",
        "lubricating oil", "润滑油", "hydraulic oil", "液压油",
    ]),
    ("Tools", [
        "combination pliers", "钳子", "diagonal pliers", "偏口钳",
        "needle nose pliers", "尖嘴钳", "adjustable wrench", "活口扳手",
        "ring spanner", "梅花扳手", "slotted screwdriver", "一字螺丝刀",
        "bolt cutter", "断线钳", "measurement tape", "钢卷尺",
        "grinder cutting wheel", "割片", "cutting blade", "切割片",
        "impact socket", "套筒", "drill bit", "钻头",
        "external snap ring pliers", "外卡簧钳", "internal snap ring pliers", "内卡簧钳",
        "pneumatic angle grinder", "风动角磨机",
        "cordless impact wrench", "cordless brushless", "电动扳手", "cordless hammer drill", "电锤钻",
        "pneumatic jack hammer", "风镐", "hydraulic jack", "千斤顶",
        "webbing sling", "吊带", "ratchet", "ratchet drive", "handle t drive",
        "ring pas", "vice grip", "wrench", "扳手", "pliers", "cutter", "drill",
        "grinder", "screwdriver", "socket", "tool bag", "工具包",
        "single wheel barrow", "手推车",
        "rachet", "ratchet", "dongkrak", "hydraulic jack", "千斤顶",
        "oring kit", "o-ring kit", "socket set",
        "plastic wrap", "透明胶", "transparent", "cling wrap",
    ]),
]


def classify_category(name_cn: str, name_en: str, spec: str = "") -> str:
    text = " ".join([name_cn, name_en, spec]).lower()
    for category, keywords in CATEGORY_RULES:
        for kw in keywords:
            if kw.lower() in text:
                return category
    return "Uncategorized"


# ─── Name / UOM parsing ───────────────────────────────────────────────────────
CN_CHAR = re.compile(r'[\u4e00-\u9fff]')


def split_bilingual(name: str) -> tuple[str, str]:
    """
    Split a bilingual string into (name_cn, name_en).

    Handles all layouts:
      'CN EN'          → ('CN', 'EN')
      'EN CN'          → ('CN', 'EN')
      'EN1 CN EN2'     → ('CN', 'EN1 EN2')   e.g. 'VRLA 电池 Lead Acid Battery'
      pure CN          → ('CN', '')
      pure EN          → ('', 'EN')
    """
    if not name:
        return "", ""
    name = str(name).strip()
    if not name:
        return "", ""

    has_cn = bool(CN_CHAR.search(name))
    has_en = bool(re.search(r'[A-Za-z]', name))

    if has_cn and not has_en:
        return name, ""
    if has_en and not has_cn:
        return "", name

    # Collect all contiguous CN runs and EN runs as separate pools.
    # This correctly handles EN-CN, CN-EN, and EN-CN-EN layouts.
    cn_runs = re.findall(r'[\u4e00-\u9fff]+(?:\s*[\u4e00-\u9fff]+)*', name)
    en_runs = re.findall(
        r"[A-Za-z][A-Za-z0-9'\"\-\.\/\\*%()（）#@!,_&+:;°Φφ]*"
        r"(?:\s+[A-Za-z0-9'\"\-\.\/\\*%()（）#@!,_&+:;°Φφ]+)*",
        name,
    )

    cn = " ".join(p.strip() for p in cn_runs if p.strip())
    en = " ".join(p.strip() for p in en_runs if p.strip())
    return cn, en


def split_dept(dept_raw: str) -> str:
    """Return English part of bilingual dept name, or CN if no English."""
    if not dept_raw:
        return ""
    _, en = split_bilingual(str(dept_raw).strip())
    return en if en else str(dept_raw).strip()


UOM_MAP = {
    "个": "pcs", "件": "pcs", "片": "pcs", "块": "pcs",
    "把": "pcs", "根": "pcs",
    "台": "unit",
    "套": "set",
    "公斤": "kg",
    "吨": "ton",
    "卷": "roll", "盘": "roll",
    "箱": "box", "盒": "box",
    "桶": "drum",
    "袋": "bag",
    "瓶": "bottle",
    "管": "tube",
    "罐": "can",
    "梱": "bundle",
    "包": "pack",
    "pcs": "pcs", "pc": "pcs",
    "set": "set", "sets": "set",
    "unit": "unit", "units": "unit",
    "kg": "kg",
    "ton": "ton", "tons": "ton",
    "roll": "roll", "rolls": "roll",
    "box": "box", "boxes": "box",
    "drum": "drum", "drums": "drum",
    "bag": "bag", "bags": "bag",
    "bottle": "bottle", "bottles": "bottle",
    "can": "can", "cans": "can",
    "bundle": "bundle", "bundles": "bundle",
    "pack": "pack", "packs": "pack",
}


def normalize_uom(uom) -> str:
    if not uom:
        return ""
    uom_str = str(uom).strip()
    return UOM_MAP.get(uom_str, UOM_MAP.get(uom_str.lower(), uom_str))


# ─── Item dataclass ───────────────────────────────────────────────────────────
class Item:
    __slots__ = ("name_cn", "name_en", "category", "spec", "uom", "department")

    def __init__(
        self,
        name_cn: str = "",
        name_en: str = "",
        category: str = "",
        spec: str = "",
        uom: str = "",
        department: str = "",
    ):
        self.name_cn = (name_cn or "").strip()
        self.name_en = (name_en or "").strip()
        self.category = category or ""
        self.spec = (str(spec) if spec is not None else "").strip()
        self.uom = normalize_uom(uom)
        self.department = (department or "").strip()

    def match_key(self) -> str:
        cn = re.sub(r"[\s\W]", "", self.name_cn.lower())
        en = re.sub(r"[\s\W]", "", self.name_en.lower())
        spec = re.sub(r"[\s\W]", "", self.spec.lower())
        return f"{cn}|{en}|{spec}"

    def name_only_key(self) -> str:
        cn = re.sub(r"[\s\W]", "", self.name_cn.lower())
        # Strip trailing disambiguation suffix e.g. "Seamless Pipe (SCH40 4inch)" → "Seamless Pipe"
        en_base = re.sub(r"\s*\([^)]*\)\s*$", "", self.name_en)
        en = re.sub(r"[\s\W]", "", en_base.lower())
        return f"{cn}|{en}"

    def is_valid(self) -> bool:
        return bool(self.name_cn or self.name_en)


# ─── File parsers ─────────────────────────────────────────────────────────────

def _find_header(rows: list, required_col: str) -> Optional[int]:
    """Return index of the first row that contains required_col."""
    for i, row in enumerate(rows[:10]):
        if row and required_col in [str(v) for v in row if v]:
            return i
    return None


def _col_idx(header: list[str], name: str) -> Optional[int]:
    for i, h in enumerate(header):
        if name in h:
            return i
    return None


def parse_january(path: Path) -> list[Item]:
    """
    1月.xlsx — each sheet is a vessel.
    Columns: 品名 | 规格型号 | 数量 | 单位 | 箱号 | 备注
    备注 = requesting department. Skip rows where 品名 is null.
    """
    wb = openpyxl.load_workbook(path)
    items: list[Item] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        hi = _find_header(rows, "品名")
        if hi is None:
            continue

        header = [str(v).strip() if v else "" for v in rows[hi]]
        idx_name = _col_idx(header, "品名")
        idx_spec = _col_idx(header, "规格型号")
        idx_unit = _col_idx(header, "单位")
        idx_dept = _col_idx(header, "备注")

        if idx_name is None:
            continue

        for row in rows[hi + 1:]:
            if not row or idx_name >= len(row) or not row[idx_name]:
                continue
            raw_name = str(row[idx_name]).strip()
            if not raw_name or raw_name in ("品名", "名称", "Name", "Item"):
                continue

            cn, en = split_bilingual(raw_name)
            spec = str(row[idx_spec]).strip() if idx_spec and idx_spec < len(row) and row[idx_spec] else ""
            uom = str(row[idx_unit]).strip() if idx_unit and idx_unit < len(row) and row[idx_unit] else ""
            dept_raw = str(row[idx_dept]).strip() if idx_dept and idx_dept < len(row) and row[idx_dept] else sheet_name
            dept = split_dept(dept_raw) if dept_raw else sheet_name

            item = Item(name_cn=cn, name_en=en, spec=spec, uom=uom, department=dept)
            if item.is_valid():
                item.category = classify_category(cn, en, spec)
                items.append(item)

    wb.close()
    return items


def parse_bakery(path: Path) -> list[Item]:
    """
    March bakery file.
    Sheet2: EN names, CN categories, UOM — full catalog (~200 rows).
    Sheet1: bilingual names 'CN EN' in column C — used to fill name_cn.
    Department = Canteen for all items.
    """
    wb = openpyxl.load_workbook(path)
    items: list[Item] = []

    # ── Sheet2: master item list ──────────────────────────────────────────────
    ws2 = wb["Sheet2"]
    rows2 = list(ws2.iter_rows(values_only=True))

    hi2 = None
    for i, row in enumerate(rows2[:5]):
        if row and any("INGREDIENTS" in str(v).upper() for v in row if v):
            hi2 = i
            break
    if hi2 is None:
        hi2 = 0

    header2 = [str(v).strip().upper() if v else "" for v in rows2[hi2]]
    idx_cat = next((i for i, h in enumerate(header2) if "CATEGORY" in h), None)
    idx_name = next((i for i, h in enumerate(header2) if "INGREDIENTS" in h or "NAME" in h), None)
    idx_uom = next((i for i, h in enumerate(header2) if "UOM" in h), None)

    s2_items: list[Item] = []
    for row in rows2[hi2 + 1:]:
        if not row:
            continue
        name_en_raw = row[idx_name] if idx_name is not None and idx_name < len(row) else None
        if not name_en_raw:
            continue
        name_en = str(name_en_raw).strip()
        if not name_en:
            continue
        uom_raw = row[idx_uom] if idx_uom is not None and idx_uom < len(row) else None
        uom = str(uom_raw).strip() if uom_raw else ""

        item = Item(name_cn="", name_en=name_en, uom=uom, department="Canteen")
        s2_items.append(item)

    # ── Sheet1: bilingual names → build cn lookup keyed by EN ────────────────
    ws1 = wb["Sheet1"]
    cn_by_en: dict[str, str] = {}
    for row in ws1.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 3 or not row[2]:
            continue
        val = str(row[2]).strip()
        if not val or not CN_CHAR.search(val):
            continue
        cn, en = split_bilingual(val)
        if cn and en:
            key = re.sub(r"[\s\W]", "", en.lower())
            cn_by_en[key] = cn

    # Apply CN names and classify
    for item in s2_items:
        en_key = re.sub(r"[\s\W]", "", item.name_en.lower())
        if en_key in cn_by_en:
            item.name_cn = cn_by_en[en_key]
        item.category = classify_category(item.name_cn, item.name_en, item.spec)
        items.append(item)

    wb.close()
    return items


def parse_april_production(path: Path) -> list[Item]:
    """
    April production plan.
    Sheet 生产: header at row 2, cols: 序号|物料类型|名称|规格型号|qty|单位
    Sheet2: dept|name|spec|qty|unit (no header row)
    Sheet1: brand|name_en|qty — WIPRO tools, department = Maintenance
    """
    wb = openpyxl.load_workbook(path)
    items: list[Item] = []

    # ── 生产 sheet ────────────────────────────────────────────────────────────
    ws = wb["生产"]
    rows = list(ws.iter_rows(values_only=True))

    hi = _find_header(rows, "名称")
    if hi is not None:
        header = [str(v).strip() if v else "" for v in rows[hi]]
        idx_dept = _col_idx(header, "物料类型")
        idx_name = _col_idx(header, "名称")
        idx_spec = _col_idx(header, "规格型号")
        idx_unit = _col_idx(header, "单位")

        for row in rows[hi + 1:]:
            if not row:
                continue
            name_raw = row[idx_name] if idx_name and idx_name < len(row) else None
            if not name_raw:
                continue
            cn, en = split_bilingual(str(name_raw))
            spec = str(row[idx_spec]).strip() if idx_spec and idx_spec < len(row) and row[idx_spec] else ""
            uom = str(row[idx_unit]).strip() if idx_unit and idx_unit < len(row) and row[idx_unit] else ""
            dept_raw = str(row[idx_dept]).strip() if idx_dept and idx_dept < len(row) and row[idx_dept] else ""
            dept = split_dept(dept_raw)

            item = Item(name_cn=cn, name_en=en, spec=spec, uom=uom, department=dept)
            if item.is_valid():
                item.category = classify_category(cn, en, spec)
                items.append(item)

    # ── Sheet2: dept|name|spec|qty|unit ──────────────────────────────────────
    ws2 = wb["Sheet2"]
    for row in ws2.iter_rows(values_only=True):
        if not row or not row[1]:
            continue
        dept_raw = str(row[0]).strip() if row[0] else ""
        name_raw = str(row[1]).strip()
        spec_raw = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        uom_raw = str(row[4]).strip() if len(row) > 4 and row[4] else ""

        cn, en = split_bilingual(name_raw)
        dept = split_dept(dept_raw)

        item = Item(name_cn=cn, name_en=en, spec=spec_raw, uom=uom_raw, department=dept)
        if item.is_valid():
            item.category = classify_category(cn, en, spec_raw)
            items.append(item)

    # ── Sheet1: WIPRO tools (brand | name_en | qty) ───────────────────────────
    ws3 = wb["Sheet1"]
    for row in ws3.iter_rows(values_only=True):
        if not row or not row[1]:
            continue
        name_en = str(row[1]).strip()
        if not name_en:
            continue
        item = Item(name_en=name_en, uom="pcs", department="Maintenance")
        item.category = classify_category("", name_en, "")
        items.append(item)

    wb.close()
    return items


def parse_generic(path: Path) -> list[Item]:
    """
    Generic parser: scan all sheets for tables with name/spec/unit columns.
    Used for future monthly files that don't match a known template.
    """
    wb = openpyxl.load_workbook(path)
    items: list[Item] = []

    NAME_KW = ["品名", "名称", "name", "物料名称", "item", "description", "ingredients"]
    SPEC_KW = ["规格", "spec", "型号", "规格型号"]
    UNIT_KW = ["单位", "unit", "uom"]
    DEPT_KW = ["部门", "dept", "department", "备注", "物料类型"]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Find header row
        hi = None
        for i, row in enumerate(rows[:10]):
            if not row:
                continue
            strs = [str(v).lower() for v in row if v]
            if any(any(kw in s for kw in NAME_KW) for s in strs):
                hi = i
                break
        if hi is None:
            continue

        header = [str(v).strip().lower() if v else "" for v in rows[hi]]

        def find_col(keywords):
            for i, h in enumerate(header):
                if any(kw in h for kw in keywords):
                    return i
            return None

        idx_name = find_col(NAME_KW)
        idx_spec = find_col(SPEC_KW)
        idx_unit = find_col(UNIT_KW)
        idx_dept = find_col(DEPT_KW)

        if idx_name is None:
            continue

        for row in rows[hi + 1:]:
            if not row or idx_name >= len(row) or not row[idx_name]:
                continue
            cn, en = split_bilingual(str(row[idx_name]))
            spec = str(row[idx_spec]).strip() if idx_spec and idx_spec < len(row) and row[idx_spec] else ""
            uom = str(row[idx_unit]).strip() if idx_unit and idx_unit < len(row) and row[idx_unit] else ""
            dept_raw = str(row[idx_dept]).strip() if idx_dept and idx_dept < len(row) and row[idx_dept] else ""
            dept = split_dept(dept_raw)

            item = Item(name_cn=cn, name_en=en, spec=spec, uom=uom, department=dept)
            if item.is_valid():
                item.category = classify_category(cn, en, spec)
                items.append(item)

    wb.close()
    return items


def detect_and_parse(path: Path) -> list[Item]:
    """Auto-detect file type and dispatch to the right parser."""
    name = path.name

    # Explicit name signals
    if "面包" in name or "食品" in name or "bakery" in name.lower():
        return parse_bakery(path)

    # Check sheets for production plan signal
    wb = openpyxl.load_workbook(path)
    sheets = wb.sheetnames
    wb.close()

    if "生产" in sheets:
        return parse_april_production(path)

    # Check for vessel-style January format
    wb2 = openpyxl.load_workbook(path)
    for sname in sheets[:4]:
        ws = wb2[sname]
        for row in ws.iter_rows(max_row=3, values_only=True):
            if row and "品名" in [str(v) for v in row if v]:
                wb2.close()
                return parse_january(path)
    wb2.close()

    return parse_generic(path)


# ─── Audit log ───────────────────────────────────────────────────────────────

def write_log(
    source_file: str,
    items_parsed: int,
    matched: int,
    new_added: int,
    new_item_ids: list[str],
    total_after: int,
    timestamp: Optional[str] = None,
) -> None:
    exists = LOG_PATH.exists()
    with open(LOG_PATH, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=LOG_COLUMNS)
        if not exists:
            writer.writeheader()
        writer.writerow({
            "timestamp": timestamp or datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
            "source_file": source_file,
            "items_parsed": items_parsed,
            "matched": matched,
            "new_added": new_added,
            "new_item_ids": " ".join(new_item_ids),
            "total_after": total_after,
        })


# ─── Item master I/O ──────────────────────────────────────────────────────────

def load_master(path: Path) -> list[dict]:
    if not path.exists():
        return []
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def save_master(records: list[dict], path: Path) -> None:
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS)
        writer.writeheader()
        writer.writerows(records)


def _next_id(records: list[dict]) -> str:
    nums = []
    for r in records:
        m = re.match(r"ITEM-(\d+)", r.get("item_id", ""))
        if m:
            nums.append(int(m.group(1)))
    return f"ITEM-{(max(nums) + 1 if nums else 1):04d}"


def _build_keys(records: list[dict]) -> list[str]:
    keys = []
    for r in records:
        cn = re.sub(r"[\s\W]", "", (r.get("name_cn") or "").lower())
        # Strip trailing disambiguation suffix e.g. "Seamless Pipe (SCH40 4inch)" → "Seamless Pipe"
        en_base = re.sub(r"\s*\([^)]*\)\s*$", "", (r.get("name_en") or ""))
        en = re.sub(r"[\s\W]", "", en_base.lower())
        keys.append(f"{cn}|{en}")
    return keys


def _find_match(item: Item, master_keys: list[str], threshold: int = 90) -> Optional[int]:
    """Return master index if a fuzzy name match >= threshold exists, else None."""
    query = item.name_only_key()
    if not query.replace("|", ""):
        return None

    # Exact match first
    if query in master_keys:
        return master_keys.index(query)

    result = process.extractOne(query, master_keys, scorer=fuzz.token_sort_ratio)
    if result and result[1] >= threshold:
        return result[2]
    return None


# ─── Core ingestion ───────────────────────────────────────────────────────────

def ingest_items(
    new_items: list[Item], master: list[dict], threshold: int = 90
) -> tuple[int, int, list[str]]:
    """
    Deduplicate new_items against master and append genuinely new entries.
    Returns (new_count, matched_count, new_item_ids).
    """
    master_keys = _build_keys(master)
    new_count = 0
    matched_count = 0
    new_item_ids: list[str] = []

    for item in new_items:
        if not item.is_valid():
            continue
        match_idx = _find_match(item, master_keys, threshold)
        if match_idx is not None:
            matched_count += 1
        else:
            item_id = _next_id(master)
            record = {
                "item_id": item_id,
                "name_cn": item.name_cn,
                "name_en": item.name_en,
                "category": item.category,
                "spec": item.spec,
                "uom": item.uom,
                "department": item.department,
            }
            master.append(record)
            # Keep keys in sync
            cn = re.sub(r"[\s\W]", "", item.name_cn.lower())
            en = re.sub(r"[\s\W]", "", item.name_en.lower())
            master_keys.append(f"{cn}|{en}")
            new_item_ids.append(item_id)
            new_count += 1

    return new_count, matched_count, new_item_ids


# ─── SQLite export ────────────────────────────────────────────────────────────

def export_sqlite(master: list[dict], db_path: Path) -> None:
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    cur.execute("DROP TABLE IF EXISTS items")
    cur.execute(
        """
        CREATE TABLE items (
            item_id   TEXT PRIMARY KEY,
            name_cn   TEXT,
            name_en   TEXT,
            category  TEXT,
            spec      TEXT,
            uom       TEXT,
            department TEXT
        )
        """
    )
    for r in master:
        cur.execute(
            "INSERT INTO items VALUES (?, ?, ?, ?, ?, ?, ?)",
            (
                r["item_id"], r["name_cn"], r["name_en"],
                r["category"], r["spec"], r["uom"], r["department"],
            ),
        )
    conn.commit()
    conn.close()


# ─── CLI entry points ─────────────────────────────────────────────────────────

def seed_database() -> None:
    print("Seeding item_master from source files...")
    master: list[dict] = []

    seed_files = [
        (Path("1月.xlsx"), parse_january),
        (Path("2026年03月面包房食品采购计划·.xlsx"), parse_bakery),
        (Path("2026年04月生产材料计划 UPDATED VERSION.xlsx"), parse_april_production),
    ]

    for path, parser in seed_files:
        if not path.exists():
            print(f"  SKIP (not found): {path.name}")
            continue
        items = parser(path)
        new, matched, new_ids = ingest_items(items, master)
        write_log(path.name, len(items), matched, new, new_ids, len(master))
        print(f"  {path.name}: {new} new, {matched} matched/skipped")

    save_master(master, CSV_PATH)
    export_sqlite(master, DB_PATH)
    print(f"\nDone. {len(master)} total items.")
    print(f"  CSV → {CSV_PATH}")
    print(f"  DB  → {DB_PATH}")
    print(f"  Log → {LOG_PATH}")


def ingest_file(path: Path) -> None:
    if not path.exists():
        sys.exit(f"Error: not found: {path}")

    master = load_master(CSV_PATH)
    print(f"Loaded {len(master)} existing items.")

    items = detect_and_parse(path)
    print(f"Parsed {len(items)} items from {path.name}.")

    new, matched, new_ids = ingest_items(items, master)

    save_master(master, CSV_PATH)
    export_sqlite(master, DB_PATH)
    write_log(path.name, len(items), matched, new, new_ids, len(master))

    print(f"\nResult:")
    print(f"  New items added : {new}")
    print(f"  Matched/skipped : {matched}")
    print(f"  Total in master : {len(master)}")
    if new_ids:
        print(f"  New IDs         : {' '.join(new_ids)}")
    print(f"  Log → {LOG_PATH}")


if __name__ == "__main__":
    if len(sys.argv) < 2 or sys.argv[1] in ("-h", "--help"):
        print(__doc__)
        sys.exit(0)

    if sys.argv[1] == "--seed":
        seed_database()
    else:
        ingest_file(Path(sys.argv[1]))
